import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime
import holidays

# 1. Configuración de la página
st.set_page_config(page_title="Conhecta - Gestión de Visitas", page_icon="🧬", layout="wide")

# 2. INYECCIÓN DE CSS AVANZADO (Fondo exacto de Conhecta, barra lateral fija y alto contraste)
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&display=swap');
        
        /* Aplicar tipografía global */
        * {
            font-family: 'Montserrat', sans-serif !important;
        }
        
        /* Imagen de fondo institucional provista desde la web de Conhecta */
        .stApp {
            background-image: url('https://www.conhecta.com.ar/static/media/Medicos.181e2e2026ed6ef17823.jpg');
            background-size: cover;
            background-position: center;
            background-attachment: fixed;
        }
        
        /* Capa de legibilidad para el contenido central */
        .block-container {
            background-color: rgba(255, 255, 255, 0.95);
            padding: 2.5rem !important;
            border-radius: 16px;
            box-shadow: 0 8px 32px rgba(11, 37, 69, 0.15);
            margin-top: 2rem;
            margin-bottom: 2rem;
            backdrop-filter: blur(6px);
        }

        /* Personalización de la barra lateral FIJA a la izquierda */
        [data-testid="stSidebar"] {
            background-color: rgba(11, 37, 69, 0.96) !important;
            box-shadow: 4px 0 15px rgba(0,0,0,0.15);
            backdrop-filter: blur(4px);
        }
        [data-testid="stSidebar"] h2, [data-testid="stSidebar"] p, [data-testid="stSidebar"] label p {
            color: #FFFFFF !important;
            font-weight: 600 !important;
        }
        
        /* Ajuste de los selectores múltiples dentro de la barra lateral (Alto Contraste) */
        div[data-baseweb="select"] {
            background-color: #ffffff !important;
            border: 2px solid #00bcbc !important;
            border-radius: 8px !important;
        }
        div[data-baseweb="popover"] div {
            color: #0b2545 !important;
        }
        span[data-baseweb="tag"] {
            background-color: #00bcbc !important;
            color: white !important;
            border-radius: 6px !important;
            font-weight: 600 !important;
        }
        
        /* Banner de Cabecera Estilo Conhecta */
        .conhecta-banner {
            background: linear-gradient(135deg, #0b2545 0%, #00bcbc 100%);
            padding: 2rem 2.5rem;
            border-radius: 12px;
            color: white;
            margin-bottom: 2.5rem;
            box-shadow: 0 4px 20px rgba(0, 188, 188, 0.25);
        }
        .conhecta-banner h1 {
            color: white !important;
            font-weight: 700;
            font-size: 2.2rem;
            margin: 0;
        }
        .conhecta-banner p {
            color: #e0f2f1 !important;
            margin-top: 0.5rem;
            font-size: 1.05rem;
            font-weight: 500;
        }
        
        /* Títulos de sección con alto contraste */
        .titulo-seccion {
            color: #0b2545 !important;
            font-weight: 700 !important;
            font-size: 1.4rem !important;
            margin-bottom: 1rem;
        }
        
        /* DISEÑO DE LA CAJA DE UPLOAD (File Uploader) */
        [data-testid="stFileUploader"] {
            background-color: #ffffff !important;
            border: 2px dashed #00bcbc !important;
            border-radius: 12px !important;
            padding: 1.5rem !important;
            box-shadow: 0 4px 12px rgba(0, 188, 188, 0.05) !important;
        }
        [data-testid="stFileUploader"] label p {
            color: #0b2545 !important;
            font-weight: 600 !important;
            font-size: 1.1rem !important;
        }
        
        /* Botón de envío de formulario (Aplicar filtros en el sidebar) */
        button[data-testid="stFormSubmitButton"] {
            background-color: #00bcbc !important;
            color: white !important;
            border-radius: 25px !important;
            border: none !important;
            font-weight: 600 !important;
            padding: 0.6rem 2rem !important;
            width: 100% !important;
            box-shadow: 0 4px 12px rgba(0, 188, 188, 0.2);
            transition: all 0.2s ease;
        }
        button[data-testid="stFormSubmitButton"]:hover {
            background-color: #ffffff !important;
            color: #0b2545 !important;
            box-shadow: 0 6px 15px rgba(255, 255, 255, 0.4);
        }

        /* Botón de Descarga del Excel */
        div.stDownloadButton > button {
            background: linear-gradient(135deg, #00bcbc 0%, #0b2545 100%) !important;
            color: white !important;
            border-radius: 25px !important;
            border: none !important;
            padding: 0.75rem 2.5rem !important;
            font-weight: 600 !important;
            box-shadow: 0 4px 15px rgba(0, 188, 188, 0.3) !important;
        }
        div.stDownloadButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(0, 188, 188, 0.4) !important;
        }
    </style>
""", unsafe_allow_html=True)

# 3. CABECERA INSTITUCIONAL ESTILO CONHECTA
st.markdown("""
    <div class="conhecta-banner">
        <h1>🧬 Portal de Auditoría Asistencial</h1>
        <p>Gestión inteligente de módulos prestacionales, control de guardias y liquidación automatizada de feriados nacionales.</p>
    </div>
""", unsafe_allow_html=True)

if 'filtros_aplicados' not in st.session_state:
    st.session_state.filtros_aplicados = False

# Listas maestras estables de opciones basadas en los reportes del sistema
MODULOS_SISTEMA = ["AO", "KETO", "SND", "BOMBA", "CONCENTRADOR", "CILINDRO", "PEDIATRICO", "COMPLEJO", "AE", "AP", "SI", "CP", "CD", "DI", "CEAO"]
PROFESIONALES_SISTEMA = ["Enfermero", "Enfermero Guardia", "Kinesiólogo", "Nutricionista", "Médico", "Fonoaudiólogo", "Terapista Ocupacional", "Psicólogo"]

# 4. BARRA LATERAL IZQUIERDA FIJA - CRITERIOS EN BLANCO
st.sidebar.markdown("## 🔍 Criterios de Selección")

with st.sidebar.form(key='formulario_filtros_laterales'):
    # CORREGIDO: Se usa st.multiselect puro dentro del bloque with st.sidebar
    modulos_seleccionados = st.multiselect("Tipos de Módulo:", MODULOS_SISTEMA, default=[])
    profesionales_seleccionados = st.multiselect("Especialidades a Auditar:", PROFESIONALES_SISTEMA, default=[])
    
    st.sidebar.markdown("---")
    boton_aceptar = st.form_submit_button(label="✅ Aplicar Filtros Operativos")
    if boton_aceptar:
        st.session_state.filtros_aplicados = True

# 5. ÁREA CENTRAL DE CARGA DE ARCHIVOS
st.markdown('<div class="titulo-seccion">📂 Carga de Datos Asistenciales</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader("Arrastrá o seleccioná el reporte mensual completo (.xlsx o .csv)", type=["xlsx", "csv"])

if uploaded_file is not None:
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, skiprows=1)
        else:
            df = pd.read_excel(uploaded_file, skiprows=1)
            
        columnas_requeridas = ['EstadoCoordinacion', 'TipoModulo', 'TipoProfesional', 'Profesional', 'Paciente', 'PlanCuidado', 'FechaInicioProF', 'FechaFinProf', 'DuracionMinutosProf']
        if not all(col in df.columns for col in columnas_requeridas):
            st.error("⚠️ Estructura incorrecta. Asegurate de subir el reporte de visitas completo con los encabezados originales del sistema.")
        else:
            # Si las listas están vacías o no se presionó el botón todavía, pedir configuración inicial
            if not st.session_state.filtros_aplicados or (len(modulos_seleccionados) == 0 and len(profesionales_seleccionados) == 0):
                st.info("💡 **Configuración inicial requerida:** Seleccioná al menos un módulo o especialidad en el panel fijo de la izquierda y presioná **'Aplicar Filtros Operativos'** para procesar los datos de este archivo.")
            else:
                # Filtro de seguridad base: Solo visitas Liberadas
                df_base = df[df['EstadoCoordinacion'] == 'Liberada'].copy()
                df_base['FechaInicioProF'] = pd.to_datetime(df_base['FechaInicioProF'], errors='coerce')
                df_base['FechaFinProf'] = pd.to_datetime(df_base['FechaFinProf'], errors='coerce')
                
                ar_holidays = holidays.Argentina()
                
                # Filtrado por las selecciones dinámicas del usuario
                df_filtered = df_base[
                    df_base['TipoModulo'].isin(modulos_seleccionados) & 
                    df_base['TipoProfesional'].isin(profesionales_seleccionados)
                ].copy()
                
                Rows_Auditoria = []
                
                for idx, row in df_filtered.iterrows():
                    paciente = row['Paciente']
                    modulo = row['TipoModulo']
                    prestacion = row['PlanCuidado']
                    profesional = row['Profesional']
                    tipo_prof = row['TipoProfesional']
                    minutos = row['DuracionMinutosProf']
                    start = row['FechaInicioProF']
                    end = row['FechaFinProf']
                    
                    horas_totales = 0.0
                    horas_feriado = 0.0
                    es_feriado_visita = "NO"
                    visitas_comunes = 0
                    visitas_feriado = 0
                    
                    if pd.isna(start): continue
                    if pd.isna(end): end = start
                    
                    rango_dias = pd.date_range(start=start.date(), end=end.date()).date
                    toca_feriado = any(d in ar_holidays for d in rango_dias)
                    if toca_feriado:
                        es_feriado_visita = "SÍ"
                    
                    if tipo_prof == "Enfermero Guardia":
                        if not pd.isna(minutos) and minutos > 0:
                            horas_totales = round((minutos / 60.0) * 2) / 2
                            
                            if toca_feriado:
                                if start.date() == end.date():
                                    if start.date() in ar_holidays:
                                        horas_feriado = horas_totales
                                else:
                                    medianoche = datetime.datetime.combine(start.date() + datetime.timedelta(days=1), datetime.time.min)
                                    min_dia1 = (medianoche - start).total_seconds() / 60.0
                                    min_dia2 = (end - medianoche).total_seconds() / 60.0
                                    
                                    min_feriado = 0.0
                                    if start.date() in ar_holidays: min_feriado += min_dia1
                                    if end.date() in ar_holidays: min_feriado += min_dia2
                                        
                                    horas_feriado = round((min_feriado / 60.0) * 2) / 2
                    else:
                        if es_feriado_visita == "SÍ":
                            visitas_feriado = 1
                        else:
                            visitas_comunes = 1
                    
                    Rows_Auditoria.append({
                        'Paciente': paciente,
                        'TipoModulo': modulo,
                        'PlanCuidado': prestacion,
                        'Profesional': profesional,
                        'TipoProfesional': tipo_prof,
                        'Visitas Comunes': visitas_comunes,
                        'Visitas Feriado': visitas_feriado,
                        'Horas Totales Guardia': horas_totales,
                        'Horas Feriado Guardia': horas_feriado
                    })
                    
                if len(Rows_Auditoria) == 0:
                    st.warning("No se encontraron registros que coincidan con la combinación exacta de filtros seleccionada.")
                else:
                    df_audit = pd.DataFrame(Rows_Auditoria)
                    
                    summary = df_audit.groupby(
                        ['Paciente', 'TipoModulo', 'PlanCuidado', 'Profesional', 'TipoProfesional']
                    ).agg({
                        'Visitas Comunes': 'sum',
                        'Visitas Feriado': 'sum',
                        'Horas Totales Guardia': 'sum',
                        'Horas Feriado Guardia': 'sum'
                    }).reset_index()
                    
                    def format_horas_texto(val):
                        if val == 0: return "-"
                        if val % 1 == 0: return f"{int(val)} hs"
                        return f"{int(val)} hs y media"
                    
                    summary_view = summary.copy()
                    summary_view['Horas Totales Guardia'] = summary_view['Horas Totales Guardia'].apply(format_horas_texto)
                    summary_view['Horas Feriado Guardia'] = summary_view['Horas Feriado Guardia'].apply(format_horas_texto)
                    
                    st.markdown('<div class="titulo-seccion">📊 Consolidado de Liquidación Mensual</div>', unsafe_allow_html=True)
                    st.dataframe(summary_view, use_container_width=True)
                    
                    # GENERACIÓN DEL EXCEL CORPORATIVO
                    output = io.BytesIO()
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Auditoría Visitas"
                    ws.views.sheetView[0].showGridLines = True
                    
                    header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
                    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    zebra_fill = PatternFill(start_color="F7F9FA", end_color="F7F9FA", fill_type="solid")
                    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    feriado_alerta_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
                    
                    thin_side = Side(style='thin', color='E0E0E0')
                    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    
                    ws['A1'] = "Reporte de Control Prestacional y Liquidación"
                    ws['A1'].font = Font(name="Arial", size=14, bold=True, color="0B2545")
                    ws['A2'] = f"Análisis automatizado bajo reglas de negocio Conhecta"
                    ws['A2'].font = Font(name="Arial", size=9, italic=True, color="555555")
                    
                    headers = [
                        "Paciente", "Módulo", "Plan de Cuidado / Prestación", 
                        "Profesional", "Tipo Profesional", "Visitas Comunes", 
                        "Visitas Feriado", "Horas Totales Guardia", "Horas Feriado Guardia"
                    ]
                    
                    for col_idx, h in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col_idx, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = thin_border
                    ws.row_dimensions[4].height = 28
                    
                    start_row = 5
                    for idx, row in summary.iterrows():
                        r_idx = start_row + idx
                        ws.cell(row=r_idx, column=1, value=row['Paciente']).alignment = Alignment(horizontal="left")
                        ws.cell(row=r_idx, column=2, value=row['TipoModulo']).alignment = Alignment(horizontal="center")
                        ws.cell(row=r_idx, column=3, value=row['PlanCuidado']).alignment = Alignment(horizontal="left")
                        ws.cell(row=r_idx, column=4, value=row['Profesional']).alignment = Alignment(horizontal="left")
                        ws.cell(row=r_idx, column=5, value=row['TipoProfesional']).alignment = Alignment(horizontal="left")
                        
                        v_com = ws.cell(row=r_idx, column=6, value=int(row['Visitas Comunes']))
                        v_fer = ws.cell(row=r_idx, column=7, value=int(row['Visitas Feriado']))
                        v_com.number_format = '#,##0'
                        v_fer.number_format = '#,##0'
                        
                        h_tot = ws.cell(row=r_idx, column=8, value=float(row['Horas Totales Guardia']))
                        h_fer = ws.cell(row=r_idx, column=9, value=float(row['Horas Feriado Guardia']))
                        h_tot.number_format = '#,##0.0'
                        h_fer.number_format = '#,##0.0'
                        
                        for c_num in [6, 7, 8, 9]:
                            ws.cell(row=r_idx, column=c_num).alignment = Alignment(horizontal="right")
                        
                        if row['Visitas Feriado'] > 0 or row['Horas Feriado Guardia'] > 0:
                            row_fill = feriado_alerta_fill
                        else:
                            row_fill = zebra_fill if idx % 2 == 0 else white_fill
                            
                        for c_idx in range(1, 10):
                            cell = ws.cell(row=r_idx, column=c_idx)
                            cell.font = Font(name="Arial", size=9)
                            cell.fill = row_fill
                            cell.border = thin_border
                            
                        ws.row_dimensions[r_idx].height = 18
                    
                    tot_row = start_row + len(summary)
                    ws.cell(row=tot_row, column=1, value="Total General").font = Font(name="Arial", size=9, bold=True, color="0B2545")
                    ws.cell(row=tot_row, column=1).border = thin_border
                    for c_empty in range(2, 6):
                        ws.cell(row=tot_row, column=c_empty).border = thin_border
                        
                    col_letters = ['F', 'G', 'H', 'I']
                    for col_let in col_letters:
                        c_idx_f = headers.index(headers[5 + col_letters.index(col_let)]) + 1
                        res_cell = ws.cell(row=tot_row, column=c_idx_f, value=f"=SUM({col_let}5:{col_let}{tot_row-1})")
                        res_cell.font = Font(name="Arial", size=9, bold=True, color="0B2545")
                        res_cell.border = thin_border
                        if col_let in ['H', 'I']:
                            res_cell.number_format = '#,##0.0'
                        else:
                            res_cell.number_format = '#,##0'
                    
                    ws.freeze_panes = "A5"
                    
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.row < 4: continue
                            if cell.value: max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
                        
                    wb.save(output)
                    processed_data = output.getvalue()
                    
                    st.markdown("---")
                    st.download_button(
                        label="📥 Descargar Reporte Institucional",
                        data=processed_data,
                        file_name="Auditoria_Conhecta_Visitas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
    except Exception as e:
        st.error(f"Error general de procesamiento: {e}")
    
